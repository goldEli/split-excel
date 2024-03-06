import tkinter as tk
from tkinter import filedialog
import xlrd
import openpyxl
import xlwt

def select_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx")])
    if file_path:
        process_excel(file_path)

def process_excel(file_path):
    file_extension = file_path.split(".")[-1]  # 获取文件后缀

    if file_extension == "xls":
        wb_xls = xlrd.open_workbook(file_path)
        sheets = wb_xls.sheet_names()
        for sheet_name in sheets:
            process_xls_sheet(wb_xls, sheet_name, f"{sheet_name}.xls")
    elif file_extension == "xlsx":
        wb_xlsx = openpyxl.load_workbook(file_path)
        sheets = wb_xlsx.sheetnames
        for sheet_name in sheets:
            process_xlsx_sheet(wb_xlsx, sheet_name, f"{sheet_name}.xlsx")
    else:
        print("不支持的文件类型！")

    print("生成完毕！")

def process_xls_sheet(wb, sheet_name, new_file_path):
    sheet = wb.sheet_by_name(sheet_name)
    new_wb = xlwt.Workbook()
    new_sheet = new_wb.add_sheet(sheet_name)

    for row_num in range(sheet.nrows):
        for col_num in range(sheet.ncols):
            new_sheet.write(row_num, col_num, sheet.cell_value(row_num, col_num))

    new_wb.save(new_file_path)

def process_xlsx_sheet(wb, sheet_name, new_file_path):
    new_wb = openpyxl.Workbook()
    new_wb.active.title = sheet_name

    for row in wb[sheet_name].iter_rows():
        new_wb[sheet_name].append([cell.value for cell in row])

    new_wb.save(new_file_path)

if __name__ == "__main__":
    select_file()
